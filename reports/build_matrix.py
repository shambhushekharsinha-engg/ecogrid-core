import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook and select active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "System Architecture Index"

# Set sheet grid lines visible
ws.views.sheetView[0].showGridLines = True

# Define premium, desaturated color palette
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Deep Classic Navy
STRIPE_FILL = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Muted Ice Blue
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Define premium typography standards
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="Arial", size=10, bold=False, color="000000")
FONT_BOLD_BODY = Font(name="Arial", size=10, bold=True, color="000000")

# Define subtle borders
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Header columns matching our multi-agent framework
headers = ["Component File Name", "Architectural Layer", "Primary Responsibility", "System Priority", "Cloud Dependency"]
ws.append(headers)

# Populate file system rows matching Shambhu's exact framework
project_structure = [
    ["app.py", "Presentation / Frontend UI", "Streamlit UI control panel interface displaying digital twin metrics and quick scenarios", "Critical", "Streamlit, Local Storage"],
    ["cloud_auditor.py", "Cloud Integration Layer", "Establishes secure data pipe connection directly with Google AI Studio Gemini API", "High", "google-genai SDK, Cloud Network"],
    ["data_aggregator.py", "Ingestion Engine Core", "Handles 64x dynamic processing of scenario data, live streaming signals, and raw CSV files", "Critical", "Local Data Streams"],
    ["main.py", "Central Orchestrator Shell", "Launches CLI runtime environment looping multi-agent logic chains deterministically", "High", "Local Memory Pipeline"],
    ["chaos_monkey.py", "Security / Adversarial Kernel", "Injects random/forced telemetry mutations representing ransomware or DDoS load spikes", "Medium", "Local State Filters"],
    ["mitigation_engine.py", "Physical Triage Engine", "Dynamically generates on-site field electrical engineering guidelines matching anomaly types", "Medium", "Local System Rules"],
    ["battery_system.py", "Stateful Hardware Logistics", "Computes multi-node discharge coefficients and chemical state-of-charge degradation curves", "High", "Mathematical Models"],
    ["crypto_ledger.py", "Security / Compliance Ledger", "Enforces append-only block logging utilizing SHA-256 validation stamps for auditing", "High", "Persistent JSON Chain"],
    ["scenarios.json", "Data Architecture Config", "Pre-configured baseline for 6 high-stress crisis scenarios spanning floods and solar drop-offs", "Medium", "Static Configuration Files"],
    ["agents/forecaster.json", "Decoupled Agent Properties", "Declares configuration parameters and safe operation thresholds for the Demand_Forecaster agent", "High", "Agent Configuration Matrix"],
    ["agents/arbitrageur.json", "Decoupled Agent Properties", "Declares pricing caps, battery drawing boundaries, and logic rules for the Grid_Arbitrageur agent", "High", "Agent Configuration Matrix"]
]

for row in project_structure:
    ws.append(row)

# Style headers cleanly
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = FONT_HEADER
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER

# Format data body rows
for row_idx in range(2, ws.max_row + 1):
    is_even = (row_idx % 2 == 0)
    current_fill = STRIPE_FILL if is_even else WHITE_FILL
    
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = FONT_BODY
        cell.fill = current_fill
        cell.border = THIN_BORDER
        
        # Smart alignments based on data types
        if col_idx in [1, 2]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col_idx == 1:
                cell.font = FONT_BOLD_BODY # Emphasize filenames
        elif col_idx in [4, 5]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Freeze header rows for professional interaction
ws.freeze_panes = "A2"

# Adjust row heights for proper spacing and alignment
ws.row_dimensions[1].height = 28
for r in range(2, ws.max_row + 1):
    ws.row_dimensions[r].height = 22

# Auto-fit column widths perfectly with added safety margins
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save finalized professional workbook asset
wb.save("EcoGrid_Architecture_Matrix.xlsx")